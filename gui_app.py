#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
发票识别器 GUI 应用程序

功能：
- 选择目录进行批量处理
- 显示处理进度
- 实时显示当前处理的文件
- 显示使用说明
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
import sys
from datetime import datetime
from entry import process_directory_to_xlsx
import json
import base64
import hashlib
import hmac


class InvoiceRecognizerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("发票识别器 v1.0.1")
        self.root.geometry("800x700")
        self.root.resizable(True, True)

        # 设置图标（如果有的话）
        try:
            self.root.iconbitmap("icon.ico")
        except:
            pass

        # 设置样式
        style = ttk.Style()
        style.theme_use('clam')

        # 初始化变量
        self.selected_directory = None
        self.processing_thread = None
        self.api_key = self.load_api_key()

        # 创建界面
        self.setup_ui()

    def setup_ui(self):
        """设置用户界面"""
        # 创建主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky="nsew")

        # 配置网格权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)

        # 创建各个区域
        self.create_api_config_section(main_frame)
        self.create_instructions(main_frame)
        self.create_controls(main_frame)
        self.create_progress_section(main_frame)
        self.create_log_section(main_frame)

    def create_api_config_section(self, parent):
        """创建API密钥配置区域"""
        # API配置框架
        api_frame = ttk.LabelFrame(parent, text="API密钥配置", padding="10")
        api_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 10))
        api_frame.columnconfigure(1, weight=1)
        
        # API密钥标签
        ttk.Label(api_frame, text="DeepSeek API Key:").grid(row=0, column=0, sticky="w", padx=(0, 10))
        
        # API密钥输入框
        self.api_key_var = tk.StringVar(value=self.api_key if self.api_key else "")
        self.api_key_entry = ttk.Entry(api_frame, textvariable=self.api_key_var, show="*", width=50)
        self.api_key_entry.grid(row=0, column=1, sticky="ew", padx=(0, 10))
        
        # 显示/隐藏按钮
        self.show_key_var = tk.BooleanVar()
        self.show_key_check = ttk.Checkbutton(api_frame, text="显示", variable=self.show_key_var, 
                                             command=self.toggle_api_key_visibility)
        self.show_key_check.grid(row=0, column=2, padx=(0, 10))
        
        # 保存按钮
        self.save_key_button = ttk.Button(api_frame, text="保存密钥", command=self.save_api_key)
        self.save_key_button.grid(row=0, column=3)
        
        # 状态标签
        self.api_status_label = ttk.Label(api_frame, text="")
        self.api_status_label.grid(row=1, column=0, columnspan=4, sticky="w", pady=(5, 0))
        
        # API密钥申请按钮
        api_help_button = ttk.Button(api_frame, text="🔑 查看API密钥申请步骤", command=self.show_api_help)
        api_help_button.grid(row=2, column=0, columnspan=4, sticky="w", pady=(10, 0))
        
        # 更新状态显示
        self.update_api_status()

    def toggle_api_key_visibility(self):
        """切换API密钥显示/隐藏"""
        if self.show_key_var.get():
            self.api_key_entry.config(show="")
        else:
            self.api_key_entry.config(show="*")

    def save_api_key(self):
        """保存API密钥"""
        api_key = self.api_key_var.get().strip()
        if not api_key:
            messagebox.showerror("错误", "请输入API密钥")
            return

        if not api_key.startswith("sk-") or len(api_key) <= 10:
            messagebox.showerror("错误", "API密钥格式不正确，应以'sk-'开头且长度足够")
            return

        try:
            self.encrypt_and_save_api_key(api_key)
            self.api_key = api_key
            self.update_api_status()
            messagebox.showinfo("成功", "API密钥已保存")
        except Exception as e:
            messagebox.showerror("错误", f"保存API密钥失败: {e}")

    def update_api_status(self):
        """更新API状态显示"""
        if self.api_key:
            self.api_status_label.config(text="✅ API密钥已配置", foreground="green")
        else:
            self.api_status_label.config(text="⚠️ 请配置API密钥", foreground="orange")

    def encrypt_and_save_api_key(self, api_key):
        """加密并保存API密钥"""
        # 使用简单的base64编码（实际项目中建议使用更强的加密）
        encoded_key = base64.b64encode(api_key.encode()).decode()

        # 创建配置目录
        config_dir = os.path.join(os.path.expanduser("~"), ".invoice_recognizer")
        os.makedirs(config_dir, exist_ok=True)

        # 保存到配置文件
        config_file = os.path.join(config_dir, "api_config.json")
        config_data = {
            "api_key": encoded_key,
            "saved_at": datetime.now().isoformat()
        }

        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump(config_data, f, ensure_ascii=False, indent=2)

    def load_api_key(self):
        """加载API密钥"""
        try:
            config_dir = os.path.join(os.path.expanduser("~"), ".invoice_recognizer")
            config_file = os.path.join(config_dir, "api_config.json")

            if os.path.exists(config_file):
                with open(config_file, 'r', encoding='utf-8') as f:
                    config_data = json.load(f)

                encoded_key = config_data.get("api_key")
                if encoded_key:
                    # 解码API密钥
                    api_key = base64.b64decode(encoded_key.encode()).decode()
                    return api_key
        except Exception as e:
            print(f"加载API密钥失败: {e}")

        return None

    def create_instructions(self, main_frame):
        """创建使用说明区域"""
        # 说明标题
        instructions_label = ttk.Label(main_frame, text="使用说明", font=("Arial", 12, "bold"))
        instructions_label.grid(row=1, column=0, columnspan=3, sticky="w", pady=(0, 5))
        
        # 说明文本
        instructions_text = """📋 功能说明：
• 智能识别PDF发票中的各项信息
• 支持批量处理多个PDF文件
• 自动生成Excel表格，包含28个字段
• 提供缓存机制，避免重复调用AI接口

⚠️ 重要提醒：
• 程序依赖DeepSeek的API接口，请先申请API KEY
• 程序会在PDF文件目录下生成缓存文件（cache_res_*.json），请勿删除，否则会重复调用AI接口
• AI识别结果可能不准确，建议人工复核重要数据
• 如果某个文件解析失败，Excel的"备注"列会显示错误信息

📁 使用步骤：
1. 配置API密钥（首次使用必需）
2. 选择包含PDF文件的目录
3. 点击"开始处理"按钮
4. 等待处理完成，查看生成的Excel文件

💡 提示：
• 首次处理文件会调用AI接口，需要网络连接
• 重复处理相同文件会使用缓存，节省费用
• 建议在稳定的网络环境下使用
• 可以随时中断和重新开始处理"""
        
        # 创建滚动文本框
        self.instructions_text = scrolledtext.ScrolledText(
            main_frame, 
            height=12, 
            width=80, 
            wrap=tk.WORD,
            font=("Arial", 9)
        )
        self.instructions_text.grid(row=2, column=0, columnspan=3, sticky="nsew", pady=(0, 10))
        self.instructions_text.insert(tk.END, instructions_text)
        self.instructions_text.config(state=tk.DISABLED)  # 设置为只读
        
    def show_api_help(self):
        """显示API密钥申请帮助"""
        help_text = """# 🔑 一、申请前准备

## 1. 注册账号
- 访问 [DeepSeek 开放平台官网](https://platform.deepseek.com/)
- 使用手机号注册/登录（未注册用户需输入验证码并设置密码）

## 2. 账户验证
- 完成邮箱/手机验证

---

# 🚀 二、官网申请步骤（推荐）

## 步骤 1：进入 API 管理页面

登录后 → 点击左侧菜单 **「API Keys」** 

## 步骤 2：创建 API Key

1. 点击 **「创建 API Key」** 按钮
2. 输入自定义名称（例如 `MyApp-Key`）
3. 点击 **「创建」** → **立即复制并保存密钥** 

> ⚠️ **关键提示：**
> 
> - **密钥仅显示一次！** 关闭页面后将无法再次查看完整 Key，务必立即保存。
> - 如密钥泄露，需立即删除并重建。

## 步骤 3：账户充值

- 进入 **「余额管理」** 页面，按需充值（最低 1 元即可启用服务）

---

# 📝 三、密钥格式说明

- **密钥格式：** `sk-xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx`
- **长度：** 通常为 32 位字符
- **前缀：** 必须以 `sk-` 开头

---

# 🔒 四、安全提醒

- 请妥善保管您的API密钥
- 不要将密钥分享给他人
- 如发现密钥泄露，请立即删除并重新创建
- 建议定期更换密钥以确保安全

---

# 💰 五、费用说明

- 按实际使用的token数量计费
- 支持按需充值，最低1元即可使用
- 可在控制台查看详细的使用记录和费用明细"""
        
        # 创建帮助窗口
        help_window = tk.Toplevel(self.root)
        help_window.title("API密钥申请指南")
        help_window.geometry("800x900")
        help_window.resizable(True, True)
        
        # 设置窗口图标和位置
        help_window.transient(self.root)
        help_window.grab_set()
        
        # 创建主框架
        help_frame = ttk.Frame(help_window, padding="10")
        help_frame.grid(row=0, column=0, sticky="nsew")
        
        # 配置网格权重
        help_window.columnconfigure(0, weight=1)
        help_window.rowconfigure(0, weight=1)
        help_frame.columnconfigure(0, weight=1)
        help_frame.rowconfigure(1, weight=1)
        
        # 标题
        title_label = ttk.Label(help_frame, text="🔑 DeepSeek API密钥申请指南", font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, sticky="w", pady=(0, 10))
        
        # 尝试使用HTML渲染
        try:
            import markdown
            from tkinterweb import HtmlFrame
            
            # 转换Markdown为HTML
            html_content = markdown.markdown(help_text, extensions=['tables', 'fenced_code', 'codehilite'])
            
            # 添加CSS样式
            styled_html = f"""
            <html>
            <head>
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        font-size: 14px;
                        line-height: 1.6;
                        margin: 20px;
                        color: #333;
                    }}
                    h1 {{
                        color: #2c3e50;
                        border-bottom: 2px solid #3498db;
                        padding-bottom: 10px;
                        margin-top: 30px;
                    }}
                    h2 {{
                        color: #34495e;
                        border-bottom: 1px solid #bdc3c7;
                        padding-bottom: 5px;
                        margin-top: 25px;
                    }}
                    h3 {{
                        color: #7f8c8d;
                        margin-top: 20px;
                    }}
                    ul, ol {{
                        margin-left: 20px;
                    }}
                    li {{
                        margin-bottom: 5px;
                    }}
                    blockquote {{
                        background-color: #f8f9fa;
                        border-left: 4px solid #3498db;
                        margin: 15px 0;
                        padding: 10px 20px;
                        font-style: italic;
                    }}
                    code {{
                        background-color: #f1f2f6;
                        padding: 2px 6px;
                        border-radius: 3px;
                        font-family: 'Courier New', monospace;
                    }}
                    pre {{
                        background-color: #f8f9fa;
                        border: 1px solid #e9ecef;
                        border-radius: 5px;
                        padding: 15px;
                        overflow-x: auto;
                    }}
                    hr {{
                        border: none;
                        border-top: 2px solid #ecf0f1;
                        margin: 30px 0;
                    }}
                    a {{
                        color: #3498db;
                        text-decoration: none;
                    }}
                    a:hover {{
                        text-decoration: underline;
                    }}
                    .highlight {{
                        background-color: #fff3cd;
                        padding: 10px;
                        border-radius: 5px;
                        border-left: 4px solid #ffc107;
                    }}
                </style>
            </head>
            <body>
                {html_content}
            </body>
            </html>
            """
            
            # 创建HTML显示区域
            help_text_widget = HtmlFrame(help_frame)
            help_text_widget.grid(row=1, column=0, sticky="nsew", pady=(0, 10))
            help_text_widget.load_html(styled_html)
            
        except ImportError:
            # 如果没有HTML渲染库，回退到文本显示
            help_text_widget = scrolledtext.ScrolledText(
                help_frame,
                height=35,
                width=80,
                wrap=tk.WORD,
                font=("Arial", 10)
            )
            help_text_widget.grid(row=1, column=0, sticky="nsew", pady=(0, 10))
            
            # 应用基本格式
            formatted_text = self.format_markdown_text(help_text)
            help_text_widget.insert(tk.END, formatted_text)
            help_text_widget.config(state=tk.DISABLED)  # 设置为只读
        
        # 按钮框架
        button_frame = ttk.Frame(help_frame)
        button_frame.grid(row=2, column=0, pady=(10, 0))
        
        # 关闭按钮
        close_button = ttk.Button(button_frame, text="关闭", command=help_window.destroy)
        close_button.pack(side=tk.LEFT, padx=(0, 10))
        
        # 打开官网按钮
        def open_website():
            import webbrowser
            webbrowser.open("https://platform.deepseek.com/")
        
        website_button = ttk.Button(button_frame, text="🌐 打开DeepSeek官网", command=open_website)
        website_button.pack(side=tk.LEFT)
        
        # 居中显示窗口
        help_window.update_idletasks()
        x = (help_window.winfo_screenwidth() // 2) - (help_window.winfo_width() // 2)
        y = (help_window.winfo_screenheight() // 2) - (help_window.winfo_height() // 2)
        help_window.geometry(f"+{x}+{y}")
        
    def format_markdown_text(self, text):
        """格式化Markdown文本为基本格式"""
        # 简单的Markdown格式转换
        formatted = text
        
        # 处理标题
        # formatted = formatted.replace('# ', '【')
        # formatted = formatted.replace('\n# ', '\n【')
        # formatted = formatted.replace('## ', '【')
        # formatted = formatted.replace('\n## ', '\n【')
        
        # # 处理粗体
        # formatted = formatted.replace('**', '【')
        # formatted = formatted.replace('**', '】')
        
        # # 处理代码块
        # formatted = formatted.replace('`', '【')
        # formatted = formatted.replace('`', '】')
        
        # # 处理引用
        # formatted = formatted.replace('> ', '    > ')
        
        # # 处理分隔线
        # formatted = formatted.replace('---', '─' * 50)
        
        return formatted

    def create_controls(self, main_frame):
        """创建控制按钮"""
        # 按钮框架
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=3, column=0, columnspan=3, pady=(0, 10))

        # 选择目录按钮
        self.select_dir_button = ttk.Button(
            button_frame,
            text="选择目录",
            command=self.select_directory,
            style="Accent.TButton"
        )
        self.select_dir_button.pack(side=tk.LEFT, padx=(0, 10))

        # 开始处理按钮
        self.process_button = ttk.Button(
            button_frame,
            text="开始处理",
            command=self.start_processing,
            state=tk.DISABLED
        )
        self.process_button.pack(side=tk.LEFT, padx=(0, 10))

        # 清空日志按钮
        self.clear_log_button = ttk.Button(
            button_frame,
            text="清空日志",
            command=self.clear_log
        )
        self.clear_log_button.pack(side=tk.LEFT)

        # 选中的目录标签
        self.selected_dir_label = ttk.Label(button_frame, text="未选择目录")
        self.selected_dir_label.pack(side=tk.RIGHT, padx=(10, 0))

    def create_progress_section(self, main_frame):
        """创建进度显示区域"""
        # 进度标题
        progress_label = ttk.Label(main_frame, text="处理进度", font=("Arial", 12, "bold"))
        progress_label.grid(row=4, column=0, columnspan=3, sticky="w", pady=(10, 5))

        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            main_frame,
            variable=self.progress_var,
            maximum=100,
            length=400
        )
        self.progress_bar.grid(row=5, column=0, columnspan=3, sticky="ew", pady=(0, 5))

        # 当前文件标签
        self.current_file_label = ttk.Label(main_frame, text="等待开始...", font=("Arial", 9))
        self.current_file_label.grid(row=6, column=0, columnspan=3, sticky="w", pady=(0, 10))

    def create_log_section(self, main_frame):
        """创建日志显示区域"""
        # 日志标题
        log_label = ttk.Label(main_frame, text="处理日志", font=("Arial", 12, "bold"))
        log_label.grid(row=7, column=0, columnspan=3, sticky="w", pady=(10, 5))

        # 日志文本框
        self.log_text = scrolledtext.ScrolledText(
            main_frame,
            height=10,
            width=80,
            wrap=tk.WORD,
            font=("Consolas", 9)
        )
        self.log_text.grid(row=8, column=0, columnspan=3, sticky="nsew", pady=(0, 10))

    def select_directory(self):
        """选择目录"""
        directory = filedialog.askdirectory(
            title="选择包含PDF文件的目录",
            initialdir=os.getcwd()
        )

        if directory:
            self.selected_directory = directory
            self.selected_dir_label.config(text=f"已选择: {os.path.basename(directory)}")
            self.process_button.config(state=tk.NORMAL)
            self.log_message(f"已选择目录: {directory}")

            # 检查目录中的PDF文件
            pdf_files = [f for f in os.listdir(directory) if f.lower().endswith('.pdf')]
            if pdf_files:
                self.log_message(f"发现 {len(pdf_files)} 个PDF文件")
            else:
                self.log_message("警告: 选择的目录中没有找到PDF文件")

    def start_processing(self):
        """开始处理文件"""
        if not self.selected_directory:
            messagebox.showerror("错误", "请先选择包含PDF文件的目录")
            return

        if not self.api_key:
            messagebox.showerror("错误", "请先配置API密钥")
            return

        # 禁用按钮
        self.select_dir_button.config(state=tk.DISABLED)
        self.process_button.config(state=tk.DISABLED)

        # 重置进度
        self.progress_var.set(0)
        self.current_file_label.config(text="正在启动...")

        # 在新线程中处理
        self.processing_thread = threading.Thread(target=self.process_files)
        self.processing_thread.daemon = True
        self.processing_thread.start()

    def process_files(self):
        """处理文件（在后台线程中运行）"""
        try:
            self.log_message("开始处理文件...")

            # 获取PDF文件列表
            pdf_files = [f for f in os.listdir(self.selected_directory) if f.lower().endswith('.pdf')]

            if not pdf_files:
                self.log_message("错误: 目录中没有找到PDF文件")
                self.root.after(0, lambda: messagebox.showerror("错误", "目录中没有找到PDF文件"))
                return

            # 重写process_directory_to_xlsx函数以支持进度回调
            self.process_with_progress(pdf_files)

        except Exception as e:
            self.log_message(f"处理过程中出现错误: {e}")
            self.root.after(0, lambda: messagebox.showerror("错误", f"处理过程中出现错误: {e}"))
        finally:
            # 恢复按钮状态
            self.root.after(0, self.enable_buttons)

    def process_with_progress(self, pdf_files):
        """带进度显示的文件处理"""
        import json
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        # 定义表头
        headers = [
            "序号", "发票代码", "发票号码", "数电发票号码", "销方识别号", "销方名称",
            "购方识别号", "购买方名称", "开票日期", "税收分类编码", "特定业务类型",
            "货物或应税劳务名称", "规格型号", "单位", "数量", "单价", "金额",
            "税率", "税额", "价税合计", "发票来源", "发票票种", "发票状态",
            "是否正数发票", "发票风险等级", "开票人", "备注"
        ]

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        if ws and hasattr(ws, 'title'):
            ws.title = "发票数据"

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            if cell and hasattr(cell, 'font'):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

        row_num = 2
        serial_number = 1

        # 处理每个文件
        for i, pdf_file in enumerate(pdf_files):
            # 更新进度
            progress = (i / len(pdf_files)) * 100
            self.root.after(0, lambda p=progress: self.progress_var.set(p))
            self.root.after(0, lambda f=pdf_file: self.current_file_label.config(text=f"正在处理: {f}"))

            if self.selected_directory:
                pdf_path = os.path.join(self.selected_directory, pdf_file)
            else:
                pdf_path = pdf_file
            self.log_message(f"处理文件 ({i + 1}/{len(pdf_files)}): {pdf_file}")

            try:
                # 导入解析函数
                from entry import parse_invoice_from_pdf

                # 临时设置API密钥
                import entry
                # original_key = entry.DEEP_SEEK_KEY
                entry.DEEP_SEEK_KEY = self.api_key

                # 解析PDF文件
                invoice_info = parse_invoice_from_pdf(pdf_path)

                # 为每个货物项目创建一行数据
                if invoice_info.items is not None:
                    for item in invoice_info.items:
                        row_data = [
                            serial_number, "", "", invoice_info.invoice_number,
                            invoice_info.seller_tax_id, invoice_info.seller_name,
                            invoice_info.buyer_tax_id, invoice_info.buyer_name,
                            invoice_info.invoice_date, invoice_info.tax_classification_code,
                            invoice_info.special_business_type, item.name, item.specification,
                            item.unit, item.quantity, item.unit_price, item.amount,
                            item.tax_rate, item.tax_amount, item.total_with_tax,
                            invoice_info.invoice_source, invoice_info.invoice_type,
                            invoice_info.invoice_status, "是" if invoice_info.is_positive_invoice else "否",
                            invoice_info.invoice_risk_level, invoice_info.issuer, invoice_info.remarks
                        ]

                        for col, value in enumerate(row_data, 1):
                            cell = ws.cell(row=row_num, column=col, value=value)

                        row_num += 1
                        serial_number += 1
                else:
                    # 如果没有货物信息，创建一行空数据
                    row_data = [
                        serial_number, "", "", invoice_info.invoice_number,
                        invoice_info.seller_tax_id, invoice_info.seller_name,
                        invoice_info.buyer_tax_id, invoice_info.buyer_name,
                        invoice_info.invoice_date, invoice_info.tax_classification_code,
                        invoice_info.special_business_type, "", "", "", "", "", "",
                        "", "", "", invoice_info.invoice_source, invoice_info.invoice_type,
                        invoice_info.invoice_status, "是" if invoice_info.is_positive_invoice else "否",
                        invoice_info.invoice_risk_level, invoice_info.issuer, invoice_info.remarks
                    ]

                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col, value=value)

                    row_num += 1
                    serial_number += 1

                self.log_message(f"✅ 成功处理: {pdf_file}")

            except Exception as e:
                error_message = f"解析失败 (文件: {pdf_file}): {str(e)}"
                self.log_message(f"❌ {error_message}")

                # 在Excel中添加错误信息行
                row_data = [serial_number] + [""] * 26 + [error_message]

                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    if cell and hasattr(cell, 'fill') and col == len(row_data):  # 备注列
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

                row_num += 1
                serial_number += 1

        # 调整列宽
        if ws and hasattr(ws, 'column_dimensions'):
            for col in range(1, len(headers) + 1):
                col_letter = chr(64 + col) if col <= 26 else chr(64 + col // 26) + chr(64 + col % 26)
                ws.column_dimensions[col_letter].width = 15

        # 保存文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"发票数据汇总_{timestamp}.xlsx"
        if self.selected_directory:
            output_path = os.path.join(self.selected_directory, output_file)
        else:
            output_path = output_file

        wb.save(output_path)

        # 完成处理
        self.root.after(0, lambda: self.progress_var.set(100))
        self.root.after(0, lambda: self.current_file_label.config(text="处理完成"))
        self.log_message(f"🎉 处理完成！共处理了 {len(pdf_files)} 个PDF文件，生成了 {row_num - 2} 行数据")
        self.log_message(f"📁 Excel文件已保存到: {output_path}")

        # 显示完成消息
        self.root.after(0, lambda: messagebox.showinfo("完成",
                                                       f"处理完成！\n\n共处理了 {len(pdf_files)} 个PDF文件\n生成了 {row_num - 2} 行数据\n\nExcel文件已保存到:\n{output_path}"))

    def enable_buttons(self):
        """恢复按钮状态"""
        self.select_dir_button.config(state=tk.NORMAL)
        self.process_button.config(state=tk.NORMAL)

    def log_message(self, message):
        """添加日志消息"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"

        self.root.after(0, lambda: self.log_text.insert(tk.END, log_entry))
        self.root.after(0, lambda: self.log_text.see(tk.END))

    def clear_log(self):
        """清空日志"""
        self.log_text.delete(1.0, tk.END)


def main():
    """主函数"""
    root = tk.Tk()
    app = InvoiceRecognizerGUI(root)

    # 设置窗口关闭事件
    def on_closing():
        # 检查是否有正在运行的线程
        if (hasattr(app, 'processing_thread') and
            app.processing_thread is not None and
            hasattr(app.processing_thread, 'is_alive') and
            app.processing_thread.is_alive()):

            if messagebox.askokcancel("退出", "正在处理文件，确定要退出吗？"):
                root.destroy()
        else:
            root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_closing)

    # 启动GUI
    root.mainloop()


if __name__ == "__main__":
    main()
