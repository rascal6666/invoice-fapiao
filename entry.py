import os
import pdfplumber
from openai import OpenAI
from dataclasses import dataclass
from typing import List, Optional, Union
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.cell import Cell
from datetime import datetime

DEEP_SEEK_KEY = ""
DEEP_SEEK_API_HOST = "https://api.deepseek.com"

SYSTEM_PROMPT = """你是一个发票识别助手，请根据描述的发票内容，识别出发票的各项信息。返回一个符合json格式的字符串。

输入格式为 : [[left,top,right,bottom,text], ...] 其中每一个元素是[left,top,right,bottom,text]。

请识别并返回以下字段的JSON格式（如果没有的话填空）：
{
  "invoice_number": "发票号码",
  "seller_tax_id": "销方识别号",
  "seller_name": "销方名称", 
  "buyer_tax_id": "购方识别号",
  "buyer_name": "购买方名称",
  "invoice_date": "开票日期",
  "tax_classification_code": "税收分类编码",
  "special_business_type": "特定业务类型",
  "items": [
    {
      "name": "货物名称",
      "specification": "规格型号",
      "unit": "单位",
      "quantity": 数量,
      "unit_price": 单价,
      "amount": 金额,
      "tax_rate": "税率",
      "tax_amount": 税额,
      "total_with_tax": 价税合计
    }
  ],
  "invoice_source": "发票来源",
  "invoice_type": "发票票种",
  "invoice_status": "发票状态",
  "is_positive_invoice": true,
  "invoice_risk_level": "发票风险等级",
  "issuer": "开票人",
  "remarks": "备注"
}

请注意：
1. 每个货物的单价使用识别到的原来单价，不要截取小数点后的位数
2. 每个货物的价税合计需要使用商品金额和税额相加得到
3. 备注要包含"备注"区域里的多个属性

例如:
输入:
```
[[161, 22, 422, 42, '电子发票（增值税专用发票）'], [438, 31, 571, 41, '发票号码：24322000000479248343'], [438, 48, 544, 58, '开票日期：2024年11月29日'], [16, 92, 24, 101, '购'], [32, 95, 209, 104, '名称：至信搏远（安徽）新材料科技有限公司'], [301, 92, 309, 101, '销'], [317, 95, 457, 104, '名称：苏州诚利恩服装科技有限公司'], [16, 102, 24, 111, '买'], [301, 102, 309, 111, '售'], [16, 112, 24, 121, '方'], [301, 112, 309, 121, '方'], [16, 122, 24, 131, '信'], [32, 125, 282, 137, '统一社会信用代码/纳税人识别号：91340700MA8P9Y7Y9D'], [301, 122, 309, 131, '信'], [317, 125, 567, 137, '统一社会信用代码/纳税人识别号：91320506MA1MMRPX1T'], [16, 132, 24, 141, '息'], [301, 132, 309, 141, '息'], [45, 151, 81, 160, '项目名称'], [119, 151, 155, 160, '规格型号'], [189, 151, 198, 160, '单'], [208, 151, 217, 160, '位'], [263, 151, 272, 160, '数'], [281, 151, 290, 160, '量'], [334, 151, 343, 160, '单'], [352, 151, 361, 160, '价'], [406, 151, 415, 160, '金'], [424, 151, 433, 160, '额'], [446, 151, 496, 160, '税率/征收率'], [551, 151, 560, 160, '税'], [569, 151, 578, 160, '额'], [12, 160, 66, 169, '*服装*净化服'], [198, 160, 207, 169, '件'], [281, 160, 290, 169, '24'], [297, 161, 361, 169, '48.6725663716814'], [402, 160, 433, 169, '1168.14'], [465, 160, 478, 169, '13%'], [555, 160, 582, 169, '151.86'], [12, 172, 57, 181, '*鞋*防砸鞋'], [198, 173, 207, 182, '双'], [281, 173, 290, 182, '18'], [297, 174, 361, 182, '64.6017699115044'], [402, 173, 433, 182, '1162.83'], [465, 173, 478, 182, '13%'], [555, 173, 582, 182, '151.17'], [58, 261, 67, 270, '合'], [103, 261, 112, 270, '计'], [397, 260, 435, 271, '¥2330.97'], [548, 260, 582, 271, '¥303.03'], [47, 280, 119, 289, '价税合计（大写）'], [178, 278, 259, 287, '贰仟陆佰叁拾肆圆整'], [406, 276, 485, 289, '（小写）¥2634.00'], [31, 296, 157, 305, '销方开户行：苏州银行浦庄支行'], [184, 296, 283, 305, '开户行号：313305060355'], [314, 296, 350, 305, '银行账号'], [355, 296, 463, 305, '：7066601841120184002636'], [31, 305, 139, 314, '订单号：IB-AH-2024102401'], [17, 309, 26, 318, '备'], [17, 326, 26, 335, '注'], [55, 367, 117, 377, '开票人：沈辰虹'], [91, 812, 121, 822, '沈辰虹']]
```
输出:
```
{
  "invoice_number": "24322000000479248343",
  "seller_tax_id": "91320506MA1MMRPX1T",
  "seller_name": "苏州诚利恩服装科技有限公司",
  "buyer_tax_id": "91340700MA8P9Y7Y9D",
  "buyer_name": "至信搏远（安徽）新材料科技有限公司",
  "invoice_date": "2024年11月29日",
  "tax_classification_code": "",
  "special_business_type": "",
  "items": [
    {
      "name": "*服装*净化服",
      "specification": "",
      "unit": "件",
      "quantity": 24,
      "unit_price": 48.6725663716814,
      "amount": 1168.14,
      "tax_rate": "13%",
      "tax_amount": 151.86,
      "total_with_tax": 1320.0
    },
    {
      "name": "*鞋*防砸鞋",
      "specification": "",
      "unit": "双",
      "quantity": 18,
      "unit_price": 64.6017699115044,
      "amount": 1162.83,
      "tax_rate": "13%",
      "tax_amount": 151.17,
      "total_with_tax": 1314.0
    }
  ],
  "invoice_source": "",
  "invoice_type": "电子发票（增值税专用发票）",
  "invoice_status": "",
  "is_positive_invoice": true,
  "invoice_risk_level": "",
  "issuer": "沈辰虹",
  "remarks": "订单号：IB-AH-2024102401, 销方开户行：苏州银行浦庄支行, 开户行号：313305060355, 银行账号：7066601841120184002636"
}
```
"""

client = OpenAI(
    api_key=DEEP_SEEK_KEY, base_url=DEEP_SEEK_API_HOST  # 替换为代理地址[5,6](@ref)
)


@dataclass
class InvoiceItem:
    """Invoice item - 发票货物项目"""

    name: str  # 货物名称
    specification: str = ""  # 规格型号
    unit: str = ""  # 单位
    quantity: Union[int, float] = 0  # 数量
    unit_price: Union[int, float] = 0.0  # 单价
    amount: Union[int, float] = 0.0  # 金额
    tax_rate: str = ""  # 税率
    tax_amount: Union[int, float] = 0.0  # 税额
    total_with_tax: Union[int, float] = 0.0  # 价税合计


@dataclass
class InvoiceInfo:
    """Invoice information data class - 发票信息数据类"""

    invoice_number: str = ""  # 数电发票号码
    seller_tax_id: str = ""  # 销方识别号
    seller_name: str = ""  # 销方名称
    buyer_tax_id: str = ""  # 购方识别号
    buyer_name: str = ""  # 购买方名称
    invoice_date: str = ""  # 开票日期
    tax_classification_code: str = ""  # 税收分类编码
    special_business_type: str = ""  # 特定业务类型
    items: Optional[List[InvoiceItem]] = None  # 货物列表
    invoice_source: str = ""  # 发票来源
    invoice_type: str = ""  # 发票票种
    invoice_status: str = ""  # 发票状态
    is_positive_invoice: bool = True  # 是否正数发票
    invoice_risk_level: str = ""  # 发票风险等级
    issuer: str = ""  # 开票人
    remarks: str = ""  # 备注

    def __post_init__(self):
        if self.items is None:
            self.items = []


"""
输入格式为 :
[[161, 22, 422, 42, '电子发票（增值税专用发票）'], [438, 31, 571, 41, '发票号码：24322000000479248343'], [438, 48, 544, 58, '开票日期：2024年11月29日'], [16, 92, 24, 101, '购'], [32, 95, 209, 104, '名称：至信搏远（安徽）新材料科技有限公司'], [301, 92, 309, 101, '销'], [317, 95, 457, 104, '名称：苏州诚利恩服装科技有限公司'], [16, 102, 24, 111, '买'], [301, 102, 309, 111, '售'], [16, 112, 24, 121, '方'], [301, 112, 309, 121, '方'], [16, 122, 24, 131, '信'], [32, 125, 282, 137, '统一社会信用代码/纳税人识别号：91340700MA8P9Y7Y9D'], [301, 122, 309, 131, '信'], [317, 125, 567, 137, '统一社会信用代码/纳税人识别号：91320506MA1MMRPX1T'], [16, 132, 24, 141, '息'], [301, 132, 309, 141, '息'], [45, 151, 81, 160, '项目名称'], [119, 151, 155, 160, '规格型号'], [189, 151, 198, 160, '单'], [208, 151, 217, 160, '位'], [263, 151, 272, 160, '数'], [281, 151, 290, 160, '量'], [334, 151, 343, 160, '单'], [352, 151, 361, 160, '价'], [406, 151, 415, 160, '金'], [424, 151, 433, 160, '额'], [446, 151, 496, 160, '税率/征收率'], [551, 151, 560, 160, '税'], [569, 151, 578, 160, '额'], [12, 160, 66, 169, '*服装*净化服'], [198, 160, 207, 169, '件'], [281, 160, 290, 169, '24'], [297, 161, 361, 169, '48.6725663716814'], [402, 160, 433, 169, '1168.14'], [465, 160, 478, 169, '13%'], [555, 160, 582, 169, '151.86'], [12, 172, 57, 181, '*鞋*防砸鞋'], [198, 173, 207, 182, '双'], [281, 173, 290, 182, '18'], [297, 174, 361, 182, '64.6017699115044'], [402, 173, 433, 182, '1162.83'], [465, 173, 478, 182, '13%'], [555, 173, 582, 182, '151.17'], [58, 261, 67, 270, '合'], [103, 261, 112, 270, '计'], [397, 260, 435, 271, '¥2330.97'], [548, 260, 582, 271, '¥303.03'], [47, 280, 119, 289, '价税合计（大写）'], [178, 278, 259, 287, '贰仟陆佰叁拾肆圆整'], [406, 276, 485, 289, '（小写）¥2634.00'], [31, 296, 157, 305, '销方开户行：苏州银行浦庄支行'], [184, 296, 283, 305, '开户行号：313305060355'], [314, 296, 350, 305, '银行账号'], [355, 296, 463, 305, '：7066601841120184002636'], [31, 305, 139, 314, '订单号：IB-AH-2024102401'], [17, 309, 26, 318, '备'], [17, 326, 26, 335, '注'], [55, 367, 117, 377, '开票人：沈辰虹'], [91, 812, 121, 822, '沈辰虹']] 
其中每一个元素是[left,top,right,bottom,text]。

要求识别出发票的各项信息，并返回一个符合json格式的字符串。

"""


def ask_deep_seek(content: str):
    response = client.chat.completions.create(
        model="deepseek-chat",
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": content},
        ],
        response_format={"type": "json_object"},
    )

    return response.choices[0].message.content


def parse_invoice_from_pdf(file_path: str) -> InvoiceInfo:
    """
    从PDF文件解析发票信息，支持缓存机制

    Args:
        file_path: PDF文件路径

    Returns:
        InvoiceInfo: 解析后的发票信息对象
    """
    # 生成缓存文件路径
    file_dir = os.path.dirname(file_path)
    file_name = os.path.basename(file_path)
    cache_file = os.path.join(file_dir, f"cache_res_{file_name}.json")
    
    # 检查缓存文件是否存在
    if os.path.exists(cache_file):
        print(f"发现缓存文件，直接读取: {cache_file}")
        try:
            with open(cache_file, 'r', encoding='utf-8') as f:
                cached_data = json.load(f)
            
            # 从缓存数据重建InvoiceInfo对象
            invoice_info = InvoiceInfo()
            invoice_info.invoice_number = cached_data.get("invoice_number", "")
            invoice_info.seller_tax_id = cached_data.get("seller_tax_id", "")
            invoice_info.seller_name = cached_data.get("seller_name", "")
            invoice_info.buyer_tax_id = cached_data.get("buyer_tax_id", "")
            invoice_info.buyer_name = cached_data.get("buyer_name", "")
            invoice_info.invoice_date = cached_data.get("invoice_date", "")
            invoice_info.tax_classification_code = cached_data.get("tax_classification_code", "")
            invoice_info.special_business_type = cached_data.get("special_business_type", "")
            invoice_info.invoice_source = cached_data.get("invoice_source", "")
            invoice_info.invoice_type = cached_data.get("invoice_type", "")
            invoice_info.invoice_status = cached_data.get("invoice_status", "")
            invoice_info.is_positive_invoice = cached_data.get("is_positive_invoice", True)
            invoice_info.invoice_risk_level = cached_data.get("invoice_risk_level", "")
            invoice_info.issuer = cached_data.get("issuer", "")
            invoice_info.remarks = cached_data.get("remarks", "")
            
            # 重建货物信息
            items_data = cached_data.get("items", [])
            for item_data in items_data:
                item = InvoiceItem(
                    name=item_data.get("name", ""),
                    specification=item_data.get("specification", ""),
                    unit=item_data.get("unit", ""),
                    quantity=item_data.get("quantity", 0),
                    unit_price=item_data.get("unit_price", 0.0),
                    amount=item_data.get("amount", 0.0),
                    tax_rate=item_data.get("tax_rate", ""),
                    tax_amount=item_data.get("tax_amount", 0.0),
                    total_with_tax=item_data.get("total_with_tax", 0.0),
                )
                assert invoice_info.items is not None
                invoice_info.items.append(item)
            
            return invoice_info
            
        except Exception as e:
            print(f"读取缓存文件失败 (文件: {os.path.basename(file_path)}): {e}，将重新解析PDF")
    
    # 如果没有缓存或缓存读取失败，则解析PDF
    print(f"开始解析PDF文件: {file_path}")
    
    # 读取PDF文件
    rs, simple = pdf_read_text(file_path)

    # 将坐标文本数据转换为字符串格式
    content = str(rs)

    # 调用AI解析发票信息
    response = ask_deep_seek(content)

    # 解析JSON响应
    try:
        if response is None:
            raise ValueError("AI响应为空")
        invoice_data = json.loads(response)

        # 创建InvoiceInfo对象
        invoice_info = InvoiceInfo()

        # 填充基本信息
        invoice_info.invoice_number = invoice_data.get("invoice_number", "")
        invoice_info.seller_tax_id = invoice_data.get("seller_tax_id", "")
        invoice_info.seller_name = invoice_data.get("seller_name", "")
        invoice_info.buyer_tax_id = invoice_data.get("buyer_tax_id", "")
        invoice_info.buyer_name = invoice_data.get("buyer_name", "")
        invoice_info.invoice_date = invoice_data.get("invoice_date", "")
        invoice_info.tax_classification_code = invoice_data.get(
            "tax_classification_code", ""
        )
        invoice_info.special_business_type = invoice_data.get(
            "special_business_type", ""
        )
        invoice_info.invoice_source = invoice_data.get("invoice_source", "")
        invoice_info.invoice_type = invoice_data.get("invoice_type", "")
        invoice_info.invoice_status = invoice_data.get("invoice_status", "")
        invoice_info.is_positive_invoice = invoice_data.get("is_positive_invoice", True)
        invoice_info.invoice_risk_level = invoice_data.get("invoice_risk_level", "")
        invoice_info.issuer = invoice_data.get("issuer", "")
        invoice_info.remarks = invoice_data.get("remarks", "")

        # 填充货物信息
        items_data = invoice_data.get("items", [])
        for item_data in items_data:
            item = InvoiceItem(
                name=item_data.get("name", ""),
                specification=item_data.get("specification", ""),
                unit=item_data.get("unit", ""),
                quantity=item_data.get("quantity", 0),
                unit_price=item_data.get("unit_price", 0.0),
                amount=item_data.get("amount", 0.0),
                tax_rate=item_data.get("tax_rate", ""),
                tax_amount=item_data.get("tax_amount", 0.0),
                total_with_tax=item_data.get("total_with_tax", 0.0),
            )
            # 确保items不为None（通过__post_init__已经初始化）
            assert invoice_info.items is not None
            invoice_info.items.append(item)

        # 保存缓存文件
        try:
            cache_data = {
                "invoice_number": invoice_info.invoice_number,
                "seller_tax_id": invoice_info.seller_tax_id,
                "seller_name": invoice_info.seller_name,
                "buyer_tax_id": invoice_info.buyer_tax_id,
                "buyer_name": invoice_info.buyer_name,
                "invoice_date": invoice_info.invoice_date,
                "tax_classification_code": invoice_info.tax_classification_code,
                "special_business_type": invoice_info.special_business_type,
                "invoice_source": invoice_info.invoice_source,
                "invoice_type": invoice_info.invoice_type,
                "invoice_status": invoice_info.invoice_status,
                "is_positive_invoice": invoice_info.is_positive_invoice,
                "invoice_risk_level": invoice_info.invoice_risk_level,
                "issuer": invoice_info.issuer,
                "remarks": invoice_info.remarks,
                "items": [
                    {
                        "name": item.name,
                        "specification": item.specification,
                        "unit": item.unit,
                        "quantity": item.quantity,
                        "unit_price": item.unit_price,
                        "amount": item.amount,
                        "tax_rate": item.tax_rate,
                        "tax_amount": item.tax_amount,
                        "total_with_tax": item.total_with_tax,
                    }
                    for item in (invoice_info.items or [])
                ]
            }
            
            with open(cache_file, 'w', encoding='utf-8') as f:
                json.dump(cache_data, f, ensure_ascii=False, indent=2)
            print(f"缓存文件已保存: {cache_file}")
            
        except Exception as e:
            print(f"保存缓存文件失败 (文件: {os.path.basename(file_path)}): {e}")

        return invoice_info

    except json.JSONDecodeError as e:
        raise ValueError(f"解析AI响应失败 (文件: {os.path.basename(file_path)}): {e}")
    except Exception as e:
        raise Exception(f"处理发票信息时出错 (文件: {os.path.basename(file_path)}): {e}")


def pdf_read_text(path):
    rs = []
    with pdfplumber.open(path) as pdf:
        page = pdf.pages[0]
        lines = page.extract_words()
        simple = page.extract_text_simple()
        simple = simple.replace(" ", "")
        simple = simple.split("\n")

        for line in lines:
            x0 = line.get("x0")
            top = line.get("top")
            x1 = line.get("x1")
            bottom = line.get("bottom")
            text = line.get("text")
            
            # 检查所有必需的值是否存在
            if x0 is not None and top is not None and x1 is not None and bottom is not None and text is not None:
                item = [
                    int(x0),
                    int(top),
                    int(x1),
                    int(bottom),
                    text,
                ]
                rs.append(item)

    return rs, simple


def process_directory_to_xlsx(
        directory_path: str, output_file: str = "invoice_data.xlsx"
):
    """
    处理目录中所有PDF文件并生成XLSX表格

    Args:
        directory_path: PDF文件所在目录路径
        output_file: 输出的XLSX文件名
    """
    # 定义表头（根据图片中的27个字段）
    headers = [
        "序号",
        "发票代码",
        "发票号码",
        "数电发票号码",
        "销方识别号",
        "销方名称",
        "购方识别号",
        "购买方名称",
        "开票日期",
        "税收分类编码",
        "特定业务类型",
        "货物或应税劳务名称",
        "规格型号",
        "单位",
        "数量",
        "单价",
        "金额",
        "税率",
        "税额",
        "价税合计",
        "发票来源",
        "发票票种",
        "发票状态",
        "是否正数发票",
        "发票风险等级",
        "开票人",
        "备注",
    ]

    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("发票数据")
    ws.title = "发票数据"

    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="808080", end_color="808080", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 获取目录中所有PDF文件
    pdf_files = [f for f in os.listdir(directory_path) if f.lower().endswith(".pdf")]

    if not pdf_files:
        print(f"在目录 {directory_path} 中未找到PDF文件")
        return

    print(f"找到 {len(pdf_files)} 个PDF文件，开始处理...")

    row_num = 2  # 从第2行开始写入数据
    serial_number = 1  # 序号计数器

    for pdf_file in pdf_files:
        pdf_path = os.path.join(directory_path, pdf_file)
        print(f"正在处理: {pdf_file}")

        try:
            # 解析PDF文件
            invoice_info = parse_invoice_from_pdf(pdf_path)

            # 为每个货物项目创建一行数据
            if invoice_info.items is not None:
                for item in invoice_info.items:
                    # 填充行数据
                    row_data = [
                        serial_number,  # 序号
                        "",  # 发票代码（通常PDF中不包含）
                        "",  # 发票号码（留空，因为数电发票号码列会填写）
                        invoice_info.invoice_number,  # 数电发票号码
                        invoice_info.seller_tax_id,  # 销方识别号
                        invoice_info.seller_name,  # 销方名称
                        invoice_info.buyer_tax_id,  # 购方识别号
                        invoice_info.buyer_name,  # 购买方名称
                        invoice_info.invoice_date,  # 开票日期
                        invoice_info.tax_classification_code,  # 税收分类编码
                        invoice_info.special_business_type,  # 特定业务类型
                        item.name,  # 货物或应税劳务名称
                        item.specification,  # 规格型号
                        item.unit,  # 单位
                        item.quantity,  # 数量
                        item.unit_price,  # 单价
                        item.amount,  # 金额
                        item.tax_rate,  # 税率
                        item.tax_amount,  # 税额
                        item.total_with_tax,  # 价税合计
                        invoice_info.invoice_source,  # 发票来源
                        invoice_info.invoice_type,  # 发票票种
                        invoice_info.invoice_status,  # 发票状态
                        "是" if invoice_info.is_positive_invoice else "否",  # 是否正数发票
                        invoice_info.invoice_risk_level,  # 发票风险等级
                        invoice_info.issuer,  # 开票人
                        invoice_info.remarks,  # 备注
                    ]

                    # 写入行数据
                    for col, value in enumerate(row_data, 1):
                        ws.cell(row=row_num, column=col, value=value)

                    row_num += 1
                    serial_number += 1
            else:
                # 如果没有货物信息，创建一行空数据
                row_data = [
                    serial_number,  # 序号
                    "",  # 发票代码
                    "",  # 发票号码
                    invoice_info.invoice_number,  # 数电发票号码
                    invoice_info.seller_tax_id,  # 销方识别号
                    invoice_info.seller_name,  # 销方名称
                    invoice_info.buyer_tax_id,  # 购方识别号
                    invoice_info.buyer_name,  # 购买方名称
                    invoice_info.invoice_date,  # 开票日期
                    invoice_info.tax_classification_code,  # 税收分类编码
                    invoice_info.special_business_type,  # 特定业务类型
                    "",  # 货物或应税劳务名称
                    "",  # 名称
                    "",  # 规格型号
                    "",  # 单位
                    "",  # 数量
                    "",  # 单价
                    "",  # 金额
                    "",  # 税率
                    "",  # 税额
                    "",  # 价税合计
                    invoice_info.invoice_source,  # 发票来源
                    invoice_info.invoice_type,  # 发票票种
                    invoice_info.invoice_status,  # 发票状态
                    "是" if invoice_info.is_positive_invoice else "否",  # 是否正数发票
                    invoice_info.invoice_risk_level,  # 发票风险等级
                    invoice_info.issuer,  # 开票人
                    invoice_info.remarks,  # 备注
                ]

                # 写入行数据
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col, value=value)

                row_num += 1
                serial_number += 1

        except Exception as e:
            print(f"处理文件 {pdf_file} 时出错: {e}")
            # 在Excel中添加错误信息行
            error_message = f"解析失败 (文件: {pdf_file}): {str(e)}"
            row_data = [
                serial_number,  # 序号
                "",  # 发票代码
                "",  # 发票号码
                "",  # 数电发票号码
                "",  # 销方识别号
                "",  # 销方名称
                "",  # 购方识别号
                "",  # 购买方名称
                "",  # 开票日期
                "",  # 税收分类编码
                "",  # 特定业务类型
                "",  # 货物或应税劳务名称
                "",  # 名称
                "",  # 规格型号
                "",  # 单位
                "",  # 数量
                "",  # 单价
                "",  # 金额
                "",  # 税率
                "",  # 税额
                "",  # 价税合计
                "",  # 发票来源
                "",  # 发票票种
                "",  # 发票状态
                "",  # 是否正数发票
                "",  # 发票风险等级
                "",  # 开票人
                error_message,  # 备注 - 显示错误信息
            ]

            # 写入错误信息行
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                # 为错误行设置红色背景
                if col == len(row_data):  # 备注列
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

            row_num += 1
            serial_number += 1
            # 即使出错也继续处理其他文件
            continue

    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[
            chr(64 + col) if col <= 26 else chr(64 + col // 26) + chr(64 + col % 26)
        ].width = 15

    # 保存文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"发票数据汇总_{timestamp}.xlsx"
    output_path = os.path.join(directory_path, output_file)
    wb.save(output_path)
    print(
        f"\n处理完成！共处理了 {len(pdf_files)} 个PDF文件，生成了 {row_num - 2} 行数据"
    )
    print(f"Excel文件已保存到: {output_path}")
